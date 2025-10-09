import os
import sys
import sqlite3
from datetime import datetime
import io
import tempfile
import traceback
import mimetypes

import customtkinter as ctk
from tkinter import ttk, messagebox, filedialog, Toplevel
from PIL import Image, UnidentifiedImageError
from fpdf import FPDF
import openpyxl

# ---------------- Helpers ----------------
def icon_label(icon, text):
    """Return the icon string if supported, else fallback to text."""
    try:
        return icon
    except Exception:
        return text

def get_resource_path(relative_path):
    """Get absolute path to resource, works for dev and for PyInstaller"""
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(base_path, relative_path)

# ---------------- Config ----------------
BASE_DIR = get_resource_path("")
ASSETS_DIR = get_resource_path("assets")
if not os.path.exists(ASSETS_DIR):
    os.makedirs(ASSETS_DIR)
LOGO_PATH = os.path.join(ASSETS_DIR, "logo.png")
DB_PATH = os.path.join(os.path.expanduser("~"), "Documents", "clinic.db")
CLINIC_NAME = "Dr. Abdulrahman Meawad"

# ---------------- Database ----------------
def db_connect():
    conn = sqlite3.connect(DB_PATH)
    try:
        conn.execute("PRAGMA foreign_keys = ON")
    except:
        pass
    return conn

def initialize_database():
    try:
        conn = db_connect()
        c = conn.cursor()
        c.execute('''
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE NOT NULL,
            password TEXT NOT NULL,
            role TEXT NOT NULL
        )''')
        c.execute('''
        CREATE TABLE IF NOT EXISTS patients (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            age INTEGER,
            gender TEXT,
            phone TEXT,
            address TEXT,
            occupation TEXT,
            diagnosis TEXT,
            prescription TEXT,
            last_visit TEXT,
            doctor TEXT,
            image BLOB
        )''')
        c.execute('''
        CREATE TABLE IF NOT EXISTS visits (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            patient_id INTEGER,
            date TEXT,
            diagnosis TEXT,
            prescription TEXT,
            doctor TEXT,
            price REAL,
            FOREIGN KEY(patient_id) REFERENCES patients(id)
        )''')
        c.execute('''
        CREATE TABLE IF NOT EXISTS patient_files (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            patient_id INTEGER,
            file_name TEXT,
            file_type TEXT,
            upload_date TEXT,
            file_data BLOB,
            FOREIGN KEY(patient_id) REFERENCES patients(id)
        )''')
        c.execute("SELECT id FROM users WHERE username='abdo'")
        if not c.fetchone():
            c.execute("INSERT INTO users (username, password, role) VALUES (?,?,?)",
                      ("abdo", "202300488", "Admin"))
            conn.commit()
        conn.close()
    except Exception as e:
        print(f"DB init error: {e}")
        traceback.print_exc()

initialize_database()

# ---------------- UI Setup ----------------
try:
    ctk.set_appearance_mode("Light")
    ctk.set_default_color_theme("blue")
except:
    pass

def pil_to_ctk_image(pil_image, size):
    try:
        return ctk.CTkImage(light_image=pil_image, dark_image=pil_image, size=size)
    except:
        return None

# ---------------- PDF Export ----------------
def save_patient_record_pdf(patient_data, visits_data, files_data=None):
    try:
        pdf = FPDF()
        pdf.add_page()

        # Header
        pdf.set_font("Helvetica", "B", 18)
        pdf.cell(0, 10, CLINIC_NAME, new_x="LMARGIN", new_y="NEXT", align="C")
        pdf.cell(0, 10, "Patient Record", new_x="LMARGIN", new_y="NEXT", align="C")
        pdf.ln(6)

        # Patient info
        pdf.set_font("Helvetica", "B", 14)
        pdf.cell(0, 8, f"Patient ID: {patient_data[0] or 'N/A'}", new_x="LMARGIN", new_y="NEXT")
        pdf.set_font("Helvetica", size=12)
        pdf.cell(0, 8, f"Name: {patient_data[1] or 'N/A'}", new_x="LMARGIN", new_y="NEXT")
        pdf.cell(0, 8, f"Age: {patient_data[2] or 'N/A'}", new_x="LMARGIN", new_y="NEXT")
        pdf.cell(0, 8, f"Gender: {patient_data[3] or 'N/A'}", new_x="LMARGIN", new_y="NEXT")
        pdf.cell(0, 8, f"Phone: {patient_data[4] or 'N/A'}", new_x="LMARGIN", new_y="NEXT")
        pdf.cell(0, 8, f"Address: {patient_data[5] or 'N/A'}", new_x="LMARGIN", new_y="NEXT")
        pdf.cell(0, 8, f"Occupation: {patient_data[6] or 'N/A'}", new_x="LMARGIN", new_y="NEXT")
        pdf.cell(0, 8, f"Last Visit: {patient_data[9] or 'N/A'}", new_x="LMARGIN", new_y="NEXT")
        pdf.cell(0, 8, f"Doctor: {patient_data[10] or 'N/A'}", new_x="LMARGIN", new_y="NEXT")

        # Visits
        pdf.ln(6)
        pdf.set_font("Helvetica", "B", 14)
        pdf.cell(0, 8, "Visit History", new_x="LMARGIN", new_y="NEXT")
        pdf.set_font("Helvetica", size=12)
        if visits_data:
            for visit in visits_data:
                pdf.ln(4)
                pdf.cell(0, 8, f"Visit ID: {visit[0] or 'N/A'}", new_x="LMARGIN", new_y="NEXT")
                pdf.cell(0, 8, f"Date: {visit[2] or 'N/A'}", new_x="LMARGIN", new_y="NEXT")
                pdf.cell(0, 8, f"Diagnosis: {visit[3] or 'N/A'}", new_x="LMARGIN", new_y="NEXT")
                pdf.cell(0, 8, f"Prescription: {visit[4] or 'N/A'}", new_x="LMARGIN", new_y="NEXT")
                pdf.cell(0, 8, f"Doctor: {visit[5] or 'N/A'}", new_x="LMARGIN", new_y="NEXT")
                price = visit[6] or 0.0
                pdf.cell(0, 8, f"Price: ${float(price):.2f}", new_x="LMARGIN", new_y="NEXT")
        else:
            pdf.cell(0, 8, "No visit history found", new_x="LMARGIN", new_y="NEXT")

        # Files
        pdf.ln(6)
        pdf.set_font("Helvetica", "B", 14)
        pdf.cell(0, 8, "Patient Files", new_x="LMARGIN", new_y="NEXT")
        pdf.set_font("Helvetica", size=12)
        if files_data:
            for fdata in files_data:
                pdf.ln(4)
                pdf.cell(0, 8, f"File Name: {fdata[0]}", new_x="LMARGIN", new_y="NEXT")
                pdf.cell(0, 8, f"File Type: {fdata[1]}", new_x="LMARGIN", new_y="NEXT")
                pdf.cell(0, 8, f"Upload Date: {fdata[2]}", new_x="LMARGIN", new_y="NEXT")
                if fdata[1]=="image" and fdata[3]:
                    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".png")
                    tmp.write(fdata[3]); tmp.close()
                    pdf.image(tmp.name, w=50)
                    os.unlink(tmp.name)
        else:
            pdf.cell(0, 8, "No files attached", new_x="LMARGIN", new_y="NEXT")

        # Save in Documents
        docs = os.path.join(os.path.expanduser("~"), "Documents")
        fname = os.path.join(
            docs,
            f"patient_record_{patient_data[1].replace(' ','_')}_{int(datetime.now().timestamp())}.pdf"
        )
        pdf.output(fname)
        return fname
    except Exception as e:
        print(f"Error saving PDF: {e}")
        traceback.print_exc()
        return None

# ---------------- Login Window ----------------
class LoginWindow(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title(f"{CLINIC_NAME} — Login")
        self.geometry("480x420")
        self.resizable(False, False)
        try:
            self.iconbitmap(os.path.join("assets","logo.ico"))
        except:
            pass
        frm = ctk.CTkFrame(self, width=440, height=380, corner_radius=12)
        frm.pack(pady=20, padx=20, fill="both", expand=True)
        if os.path.exists(LOGO_PATH):
            try:
                img = Image.open(LOGO_PATH); img.thumbnail((160,160))
                logo = pil_to_ctk_image(img,(160,160))
                lbl = ctk.CTkLabel(frm,image=logo,text=""); lbl.image=logo; lbl.place(x=140,y=30)
            except:
                ctk.CTkLabel(frm,text="CLINIC",font=ctk.CTkFont(size=20,weight="bold")).place(x=180,y=30)
        else:
            ctk.CTkLabel(frm,text="CLINIC",font=ctk.CTkFont(size=20,weight="bold")).place(x=180,y=30)
        ctk.CTkLabel(frm,text="WELCOME",font=ctk.CTkFont(size=30,weight="bold")).place(x=30,y=120)
        self.username=ctk.CTkEntry(frm,width=360,placeholder_text="Username");self.username.place(x=30,y=195)
        self.password=ctk.CTkEntry(frm,width=360,placeholder_text="Password",show="*");self.password.place(x=30,y=255)
        ctk.CTkButton(frm,text="Login",width=150,command=self.do_login,fg_color="#3182ce").place(x=60,y=310)
        ctk.CTkButton(frm,text="Exit",width=150,fg_color="#718096",hover_color="#4a5568",command=self.destroy).place(x=230,y=310)
        ctk.CTkLabel(frm,text="Default admin: abdo / 202300488",font=ctk.CTkFont(size=10)).place(x=30,y=350)

    def do_login(self):
        user=self.username.get().strip(); pwd=self.password.get().strip()
        if not user or not pwd:
            messagebox.showerror("Login Failed","Enter both username and password");return
        conn=db_connect(); c=conn.cursor()
        c.execute("SELECT id,username,role FROM users WHERE username=? AND password=?", (user,pwd))
        row=c.fetchone(); conn.close()
        if not row:
            messagebox.showerror("Login Failed","Invalid credentials");return
        self.destroy()
        ClinicApp({"id":row[0],"username":row[1],"role":row[2]}).mainloop()

# ---------------- Main Application ----------------
class ClinicApp(ctk.CTk):
    def __init__(self,current_user):
        super().__init__()
        self.current_user=current_user
        self.title(f"{CLINIC_NAME} — Dashboard ({current_user['username']} - {current_user['role']})")
        try:
            self.iconbitmap(os.path.join("assets","logo.ico"))
        except:
            pass
        try:
            self.state("zoomed")
        except:
            w,h=self.winfo_screenwidth(),self.winfo_screenheight()
            self.geometry(f"{w}x{h}")
        header=ctk.CTkFrame(self,fg_color="#1a3c6c",height=100); header.pack(fill="x",padx=10,pady=10)
        ctk.CTkLabel(header,text=CLINIC_NAME,font=ctk.CTkFont(size=24,weight="bold"),text_color="white").pack(pady=15)
        ctk.CTkLabel(header,text="Patient Management System",font=ctk.CTkFont(size=14),text_color="white").pack()
        ctk.CTkLabel(header,text=f"User: {current_user['username']} ({current_user['role']})",font=ctk.CTkFont(size=12),text_color="white").pack(pady=5)
        nav=ctk.CTkFrame(self,fg_color="#2c5282",height=60); nav.pack(fill="x",padx=10)
        ctk.CTkButton(nav,text="Manage Patients",command=self.open_patients,fg_color="#3182ce").pack(side="left",padx=10,pady=10)
        ctk.CTkButton(nav,text="Visit History",command=self.open_visits,fg_color="#319795").pack(side="left",padx=10,pady=10)
        if current_user['role']=="Admin":
            ctk.CTkButton(nav,text="Manage Users",command=self.open_users,fg_color="#38a169").pack(side="left",padx=10,pady=10)
        ctk.CTkButton(nav,text="Export Excel",command=self.export_patients_excel,fg_color="#dd6b20").pack(side="left",padx=10,pady=10)
        ctk.CTkButton(nav,text="Logout",command=self.logout,fg_color="#e53e3e").pack(side="right",padx=10,pady=10)
        self.content=ctk.CTkFrame(self,fg_color="#f0f4f8"); self.content.pack(fill="both",expand=True,padx=10,pady=(0,10))
        self.open_patients()

    def clear_content(self):
        for w in self.content.winfo_children(): w.destroy()

    def open_patients(self):
        self.clear_content(); PatientsView(self.content)

    def open_visits(self):
        self.clear_content(); VisitsView(self.content)

    def open_users(self):
        if self.current_user['role']!="Admin":
            messagebox.showerror("Permission denied","Admin only");return
        self.clear_content(); UsersView(self.content)

    def logout(self):
        self.destroy(); LoginWindow().mainloop()

    def export_patients_excel(self):
        path=filedialog.asksaveasfilename(defaultextension=".xlsx",filetypes=[("Excel files","*.xlsx")])
        if not path: return
        wb=openpyxl.Workbook(); ws=wb.active; ws.title="Patients"
        headers=["ID","Name","Age","Gender","Phone","Address","Occupation","Diagnosis","Prescription","Last Visit","Doctor"]
        ws.append(headers)
        conn=db_connect(); c=conn.cursor()
        c.execute("SELECT id,name,age,gender,phone,address,occupation,diagnosis,prescription,last_visit,doctor FROM patients")
        for r in c.fetchall(): ws.append([cell or "" for cell in r])
        wb.save(path); messagebox.showinfo("Exported",f"Exported {ws.max_row-1} patients to:\n{path}")

# ---------------- Patients View ----------------
class PatientsView:
    def __init__(self, parent):
        self.parent = parent
        self.current_image_blob = None
        self.patient_files = []

        parent.grid_columnconfigure(0, weight=1)
        parent.grid_columnconfigure(1, weight=2)
        parent.grid_rowconfigure(0, weight=1)

        left = ctk.CTkScrollableFrame(parent, corner_radius=8, fg_color="#e2e8f0")
        left.grid(row=0, column=0, padx=10, pady=10, sticky="nsew")

        if os.path.exists(LOGO_PATH):
            try:
                pil_logo = Image.open(LOGO_PATH)
                pil_logo.thumbnail((100, 100))
                logo_img = pil_to_ctk_image(pil_logo, size=(100, 100))
                if logo_img:
                    logo_label = ctk.CTkLabel(left, image=logo_img, text="")
                    logo_label.image = logo_img
                    logo_label.pack(pady=10)
                else:
                    ctk.CTkLabel(left, text="CLINIC LOGO",
                                font=ctk.CTkFont(size=16, weight="bold")).pack(pady=10)
            except Exception as e:
                print(f"Error loading logo: {e}")
                ctk.CTkLabel(left, text="CLINIC LOGO",
                            font=ctk.CTkFont(size=16, weight="bold")).pack(pady=10)
        else:
            ctk.CTkLabel(left, text="CLINIC LOGO",
                        font=ctk.CTkFont(size=16, weight="bold")).pack(pady=10)

        ctk.CTkLabel(left, text="Patient Information",
                    font=ctk.CTkFont(size=18, weight="bold")).pack(pady=10)

        ctk.CTkButton(left, text=icon_label("🖼️ Upload Patient Photo", "[Photo] Upload Patient Photo"),
                     command=self.upload_photo,
                     fg_color="#27ae60", hover_color="#229954").pack(pady=10, padx=20, fill="x")

        ctk.CTkButton(left, text=icon_label("📎 Upload Patient Files", "[Files] Upload Patient Files"),
                     command=self.upload_files,
                     fg_color="#3498db", hover_color="#2980b9").pack(pady=5, padx=20, fill="x")

        photo_frame = ctk.CTkFrame(left, width=160, height=160, corner_radius=8)
        photo_frame.pack(pady=10)
        photo_frame.pack_propagate(False)

        self.photo_label = ctk.CTkLabel(photo_frame, text="No Photo", font=ctk.CTkFont(size=12))
        self.photo_label.pack(expand=True)

        form_frame = ctk.CTkFrame(left, corner_radius=8, fg_color="transparent")
        form_frame.pack(fill="both", expand=True, padx=10, pady=10)

        ctk.CTkLabel(form_frame, text="1. Patient ID:",
                    font=ctk.CTkFont(size=12, weight="bold")).pack(anchor="w", padx=10, pady=(10, 0))
        self.e_id = ctk.CTkEntry(form_frame, placeholder_text="Enter ID to load/update/delete")
        self.e_id.pack(fill="x", padx=10, pady=(0, 10))

        ctk.CTkLabel(form_frame, text="2. Gender:",
                    font=ctk.CTkFont(size=12, weight="bold")).pack(anchor="w", padx=10, pady=(5, 0))
        self.gender_cb = ctk.CTkComboBox(form_frame, values=["Male", "Female", "Other"])
        self.gender_cb.set("Male")
        self.gender_cb.pack(fill="x", padx=10, pady=(0, 10))

        ctk.CTkLabel(form_frame, text="3. Full Name*:",
                    font=ctk.CTkFont(size=12, weight="bold")).pack(anchor="w", padx=10, pady=(5, 0))
        self.e_name = ctk.CTkEntry(form_frame, placeholder_text="Enter patient's full name")
        self.e_name.pack(fill="x", padx=10, pady=(0, 10))

        ctk.CTkLabel(form_frame, text="4. Age:",
                    font=ctk.CTkFont(size=12, weight="bold")).pack(anchor="w", padx=10, pady=(5, 0))
        self.e_age = ctk.CTkEntry(form_frame, placeholder_text="Enter age")
        self.e_age.pack(fill="x", padx=10, pady=(0, 10))

        ctk.CTkLabel(form_frame, text="5. Occupation:",
                    font=ctk.CTkFont(size=12, weight="bold")).pack(anchor="w", padx=10, pady=(5, 0))
        self.e_occupation = ctk.CTkEntry(form_frame, placeholder_text="Enter occupation")
        self.e_occupation.pack(fill="x", padx=10, pady=(0, 10))

        ctk.CTkLabel(form_frame, text="6. Phone:",
                    font=ctk.CTkFont(size=12, weight="bold")).pack(anchor="w", padx=10, pady=(5, 0))
        self.e_phone = ctk.CTkEntry(form_frame, placeholder_text="Enter phone number")
        self.e_phone.pack(fill="x", padx=10, pady=(0, 10))

        ctk.CTkLabel(form_frame, text="7. Diagnosis:",
                    font=ctk.CTkFont(size=12, weight="bold")).pack(anchor="w", padx=10, pady=(5, 0))
        self.e_diag = ctk.CTkEntry(form_frame, placeholder_text="Enter diagnosis")
        self.e_diag.pack(fill="x", padx=10, pady=(0, 10))

        ctk.CTkLabel(form_frame, text="8. Prescription/Notes:",
                    font=ctk.CTkFont(size=12, weight="bold")).pack(anchor="w", padx=10, pady=(5, 0))
        self.e_presc = ctk.CTkEntry(form_frame, placeholder_text="Enter prescription or notes")
        self.e_presc.pack(fill="x", padx=10, pady=(0, 10))

        ctk.CTkLabel(form_frame, text="9. Doctor:",
                    font=ctk.CTkFont(size=12, weight="bold")).pack(anchor="w", padx=10, pady=(5, 0))
        self.e_doctor = ctk.CTkEntry(form_frame, placeholder_text="Enter doctor's name")
        self.e_doctor.pack(fill="x", padx=10, pady=(0, 10))

        ctk.CTkLabel(form_frame, text="10. Address:",
                    font=ctk.CTkFont(size=12, weight="bold")).pack(anchor="w", padx=10, pady=(5, 0))
        self.e_address = ctk.CTkEntry(form_frame, placeholder_text="Enter address")
        self.e_address.pack(fill="x", padx=10, pady=(0, 10))

        ctk.CTkLabel(form_frame, text="Actions:",
                    font=ctk.CTkFont(size=14, weight="bold")).pack(anchor="w", padx=10, pady=(20, 10))

        ctk.CTkButton(form_frame, text=icon_label("➕ Add Patient", "[+] Add Patient"), command=self.add_patient,
                     fg_color="#27ae60", hover_color="#229954", height=40).pack(fill="x", padx=10, pady=5)

        ctk.CTkButton(form_frame, text=icon_label("📂 Load Patient", "[Open] Load Patient"), command=self.load_patient_by_id,
                     fg_color="#3498db", hover_color="#2980b9", height=40).pack(fill="x", padx=10, pady=5)

        ctk.CTkButton(form_frame, text=icon_label("✏️ Update Patient", "[Edit] Update Patient"), command=self.update_patient,
                     fg_color="#f39c12", hover_color="#e67e22", height=40).pack(fill="x", padx=10, pady=5)

        ctk.CTkButton(form_frame, text=icon_label("🗑️ Delete Patient", "[Del] Delete Patient"), command=self.delete_patient,
                     fg_color="#e74c3c", hover_color="#c0392b", height=40).pack(fill="x", padx=10, pady=5)

        ctk.CTkButton(form_frame, text=icon_label("📄 Export to PDF", "[PDF] Export to PDF"), command=self.export_patient_pdf,
                     fg_color="#9b59b6", hover_color="#8e44ad", height=40).pack(fill="x", padx=10, pady=5)

        ctk.CTkButton(form_frame, text=icon_label("🔄 Clear Form", "[R] Clear Form"), command=self.clear_form,
                     fg_color="#7f8c8d", hover_color="#95a5a6", height=40).pack(fill="x", padx=10, pady=5)

        right = ctk.CTkFrame(parent, corner_radius=8, fg_color="#e2e8f0")
        right.grid(row=0, column=1, padx=10, pady=10, sticky="nsew")
        right.grid_columnconfigure(0, weight=1)
        right.grid_rowconfigure(1, weight=1)

        ctk.CTkLabel(right, text="Patient Records",
                    font=ctk.CTkFont(size=18, weight="bold")).pack(pady=10)

        search_frame = ctk.CTkFrame(right, fg_color="transparent")
        search_frame.pack(fill="x", padx=10, pady=5)

        self.search = ctk.CTkEntry(search_frame, placeholder_text="Search by name, phone or doctor")
        self.search.pack(side="left", fill="x", expand=True, padx=(0, 5))

        ctk.CTkButton(search_frame, text=icon_label("🔍 Search", "[?] Search"), width=100,
                     command=self.search_patients).pack(side="left", padx=5)
        ctk.CTkButton(search_frame, text=icon_label("🔄 Refresh", "[R] Refresh"), width=100,
                     command=self.load_all_patients).pack(side="left", padx=5)

        table_frame = ctk.CTkFrame(right, fg_color="transparent")
        table_frame.pack(fill="both", expand=True, padx=10, pady=10)
        table_frame.grid_columnconfigure(0, weight=1)
        table_frame.grid_rowconfigure(0, weight=1)

        cols = ("id", "name", "age", "gender", "phone", "occupation", "doctor", "last_visit")
        self.tree = ttk.Treeview(table_frame, columns=cols, show="headings", height=20)

        self.tree.heading("id", text="ID")
        self.tree.column("id", width=50, anchor="center")

        self.tree.heading("name", text="Name")
        self.tree.column("name", width=150, anchor="w")

        self.tree.heading("age", text="Age")
        self.tree.column("age", width=50, anchor="center")

        self.tree.heading("gender", text="Gender")
        self.tree.column("gender", width=80, anchor="center")

        self.tree.heading("phone", text="Phone")
        self.tree.column("phone", width=120, anchor="w")

        self.tree.heading("occupation", text="Occupation")
        self.tree.column("occupation", width=120, anchor="w")

        self.tree.heading("doctor", text="Doctor")
        self.tree.column("doctor", width=120, anchor="w")

        self.tree.heading("last_visit", text="Last Visit")
        self.tree.column("last_visit", width=120, anchor="center")

        v_scrollbar = ctk.CTkScrollbar(table_frame, orientation="vertical", command=self.tree.yview)
        v_scrollbar.grid(row=0, column=1, sticky="ns")
        self.tree.configure(yscrollcommand=v_scrollbar.set)

        h_scrollbar = ctk.CTkScrollbar(table_frame, orientation="horizontal", command=self.tree.xview)
        h_scrollbar.grid(row=1, column=0, sticky="ew")
        self.tree.configure(xscrollcommand=h_scrollbar.set)

        self.tree.grid(row=0, column=0, sticky="nsew", padx=(0, 5))

        self.tree.bind("<Double-1>", self.on_double)

        self.load_all_patients()

    def upload_photo(self):
        try:
            path = filedialog.askopenfilename(title="Select patient photo",
                                             filetypes=[("Image files","*.png *.jpg *.jpeg *.bmp")])
            if not path:
                return
            with open(path, "rb") as f:
                blob = f.read()
            self.current_image_blob = blob
            pil_img = Image.open(io.BytesIO(blob))
            pil_img.thumbnail((160, 160))
            ctk_img = pil_to_ctk_image(pil_img, size=(160, 160))
            if ctk_img:
                self.photo_label.configure(image=ctk_img, text="")
                self.photo_label.image = ctk_img
            else:
                self.photo_label.configure(text="Photo Loaded")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load image: {e}")

    def upload_files(self):
        try:
            paths = filedialog.askopenfilenames(
                title="Select patient files",
                filetypes=[("All files","*.*"),
                           ("Image files","*.png *.jpg *.jpeg *.bmp *.gif"),
                           ("Document files","*.pdf *.doc *.docx *.txt")]
            )
            if not paths:
                return

            MAX_BYTES = 8 * 1024 * 1024  # 8 MB per file
            self.patient_files = []

            for path in paths:
                try:
                    size = os.path.getsize(path)
                except Exception:
                    size = 0
                if size > MAX_BYTES:
                    messagebox.showwarning("File skipped", f"{os.path.basename(path)} is larger than 8MB and was skipped.")
                    continue

                with open(path, "rb") as f:
                    blob = f.read()

                # Use simple categories to match existing PDF condition, or switch to MIME below
                ext = os.path.splitext(path)[1].lower()
                if ext in [".png", ".jpg", ".jpeg", ".bmp", ".gif"]:
                    ftype = "image"
                elif ext in [".pdf", ".doc", ".docx", ".txt"]:
                    ftype = "document"
                else:
                    # Optional: use mimetypes to be more precise
                    mime, _ = mimetypes.guess_type(path)
                    ftype = "image" if (mime and mime.startswith("image")) else "other"

                self.patient_files.append({
                    "name": os.path.basename(path),
                    "type": ftype,
                    "data": blob
                })

            if self.patient_files:
                messagebox.showinfo("Success", f"Queued {len(self.patient_files)} file(s) to attach to this patient")
            else:
                messagebox.showwarning("No files", "No files were added (all may have been skipped).")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to upload files: {e}")

    def add_patient(self):
        try:
            name = self.e_name.get().strip()
            if not name:
                messagebox.showerror("Error", "Name is required")
                return

            age_str = self.e_age.get().strip()
            age = None
            if age_str:
                try:
                    age = int(age_str)
                    if age < 0 or age > 150:
                        messagebox.showerror("Error", "Age must be between 0 and 150")
                        return
                except ValueError:
                    messagebox.showerror("Error", "Age must be a number")
                    return

            gender = self.gender_cb.get()
            phone = self.e_phone.get().strip()
            address = self.e_address.get().strip()
            occupation = self.e_occupation.get().strip()
            diag = self.e_diag.get().strip()
            presc = self.e_presc.get().strip()
            doctor = self.e_doctor.get().strip()
            last_visit = datetime.now().strftime("%Y-%m-%d %H:%M")

            conn = db_connect()
            c = conn.cursor()

            c.execute('''INSERT INTO patients (name, age, gender, phone, address, occupation, diagnosis, prescription, last_visit, doctor, image)
                         VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
                      (name, age, gender, phone, address, occupation, diag, presc, last_visit, doctor,
                       sqlite3.Binary(self.current_image_blob) if self.current_image_blob else None))
            conn.commit()

            patient_id = c.lastrowid

            if self.patient_files:
                for file_info in self.patient_files:
                    c.execute('''INSERT INTO patient_files (patient_id, file_name, file_type, upload_date, file_data)
                                 VALUES (?, ?, ?, ?, ?)''',
                              (patient_id, file_info["name"], file_info["type"],
                               datetime.now().strftime("%Y-%m-%d %H:%M"),
                               sqlite3.Binary(file_info["data"])))
                conn.commit()
                self.patient_files = []  # clear queued files after successful save

            conn.close()
            messagebox.showinfo("Success", "Patient added successfully")
            self.clear_form()
            self.load_all_patients()
        except Exception as e:
            messagebox.showerror("Error", f"Failed to add patient: {e}")

    def load_patient_by_id(self):
        try:
            pid = self.e_id.get().strip()
            if not pid:
                messagebox.showerror("Error", "Enter ID to load")
                return

            pid_int = int(pid)

            conn = db_connect()
            c = conn.cursor()

            c.execute("SELECT id, name, age, gender, phone, address, occupation, diagnosis, prescription, last_visit, doctor, image FROM patients WHERE id=?", (pid_int,))
            row = c.fetchone()
            if not row:
                conn.close()
                messagebox.showerror("Error", "Patient not found")
                return

            (pid, name, age, gender, phone, address, occupation, diag, presc, last_visit, doctor, image_blob) = row

            self.e_id.delete(0, "end")
            self.e_id.insert(0, str(pid))

            self.e_name.delete(0, "end")
            self.e_name.insert(0, name or "")

            self.e_age.delete(0, "end")
            self.e_age.insert(0, str(age) if age is not None else "")

            self.gender_cb.set(gender or "Male")

            self.e_phone.delete(0, "end")
            self.e_phone.insert(0, phone or "")

            self.e_address.delete(0, "end")
            self.e_address.insert(0, address or "")

            self.e_occupation.delete(0, "end")
            self.e_occupation.insert(0, occupation or "")

            self.e_diag.delete(0, "end")
            self.e_diag.insert(0, diag or "")

            self.e_presc.delete(0, "end")
            self.e_presc.insert(0, presc or "")

            self.e_doctor.delete(0, "end")
            self.e_doctor.insert(0, doctor or "")

            self.current_image_blob = image_blob
            if image_blob:
                try:
                    pil_img = Image.open(io.BytesIO(image_blob))
                    pil_img.thumbnail((160, 160))
                    ctk_img = pil_to_ctk_image(pil_img, size=(160, 160))
                    if ctk_img:
                        self.photo_label.configure(image=ctk_img, text="")
                        self.photo_label.image = ctk_img
                    else:
                        self.photo_label.configure(text="Photo")
                except Exception:
                    self.photo_label.configure(text="Invalid Image")
            else:
                self.photo_label.configure(image=None, text="No Photo")

            conn.close()
        except ValueError:
            messagebox.showerror("Error", "ID must be a number")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load patient: {e}")

    def update_patient(self):
        try:
            pid = self.e_id.get().strip()
            if not pid:
                messagebox.showerror("Error", "Enter ID to update")
                return

            pid_int = int(pid)

            name = self.e_name.get().strip()
            if not name:
                messagebox.showerror("Error", "Name is required")
                return

            age_str = self.e_age.get().strip()
            age = None
            if age_str:
                try:
                    age = int(age_str)
                    if age < 0 or age > 150:
                        messagebox.showerror("Error", "Age must be between 0 and 150")
                        return
                except ValueError:
                    messagebox.showerror("Error", "Age must be a number")
                    return

            gender = self.gender_cb.get()
            phone = self.e_phone.get().strip()
            address = self.e_address.get().strip()
            occupation = self.e_occupation.get().strip()
            diag = self.e_diag.get().strip()
            presc = self.e_presc.get().strip()
            doctor = self.e_doctor.get().strip()

            conn = db_connect()
            c = conn.cursor()

            c.execute("SELECT id FROM patients WHERE id=?", (pid_int,))
            if not c.fetchone():
                conn.close()
                messagebox.showerror("Error", "Patient not found")
                return

            c.execute('''UPDATE patients SET name=?, age=?, gender=?, phone=?, address=?, occupation=?, diagnosis=?, prescription=?, doctor=? WHERE id=?''',
                      (name, age, gender, phone, address, occupation, diag, presc, doctor, pid_int))

            if self.current_image_blob is not None:
                c.execute("UPDATE patients SET image=? WHERE id=?", (sqlite3.Binary(self.current_image_blob), pid_int))

            if self.patient_files:
                for file_info in self.patient_files:
                    c.execute('''INSERT INTO patient_files (patient_id, file_name, file_type, upload_date, file_data)
                                 VALUES (?, ?, ?, ?, ?)''',
                              (pid_int, file_info["name"], file_info["type"],
                               datetime.now().strftime("%Y-%m-%d %H:%M"),
                               sqlite3.Binary(file_info["data"])))
                self.patient_files = []

            conn.commit()
            conn.close()
            messagebox.showinfo("Success", "Patient updated successfully")
            self.clear_form()
            self.load_all_patients()
        except ValueError:
            messagebox.showerror("Error", "ID must be a number")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to update patient: {e}")

    def delete_patient(self):
        try:
            pid = self.e_id.get().strip()
            if not pid:
                messagebox.showerror("Error", "Enter ID to delete")
                return

            pid_int = int(pid)

            conn = db_connect()
            c = conn.cursor()

            c.execute("SELECT name FROM patients WHERE id=?", (pid_int,))
            row = c.fetchone()
            if not row:
                conn.close()
                messagebox.showerror("Error", "Patient not found")
                return

            patient_name = row[0]

            if messagebox.askyesno("Confirm Delete", f"Are you sure you want to delete patient '{patient_name}'?\nThis will also remove all their visits and files."):
                c.execute("DELETE FROM visits WHERE patient_id=?", (pid_int,))
                c.execute("DELETE FROM patient_files WHERE patient_id=?", (pid_int,))
                c.execute("DELETE FROM patients WHERE id=?", (pid_int,))
                conn.commit()
                conn.close()
                messagebox.showinfo("Success", "Patient deleted successfully")
                self.clear_form()
                self.load_all_patients()
        except ValueError:
            messagebox.showerror("Error", "ID must be a number")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to delete patient: {e}")

    def export_patient_pdf(self):
        try:
            pid = self.e_id.get().strip()
            if not pid:
                messagebox.showerror("Error", "Enter ID to export")
                return

            pid_int = int(pid)

            conn = db_connect()
            c = conn.cursor()

            c.execute("SELECT id, name, age, gender, phone, address, occupation, diagnosis, prescription, last_visit, doctor, image FROM patients WHERE id=?", (pid_int,))
            patient_row = c.fetchone()
            if not patient_row:
                conn.close()
                messagebox.showerror("Error", "Patient not found")
                return

            c.execute("SELECT id, patient_id, date, diagnosis, prescription, doctor, price FROM visits WHERE patient_id=? ORDER BY date DESC", (pid_int,))
            visits_rows = c.fetchall()

            c.execute("SELECT file_name, file_type, upload_date, file_data FROM patient_files WHERE patient_id=? ORDER BY upload_date DESC", (pid_int,))
            files_rows = c.fetchall()

            conn.close()

            pdf_path = save_patient_record_pdf(patient_row, visits_rows, files_rows)
            if pdf_path and os.path.exists(pdf_path):
                messagebox.showinfo("Success", f"Patient record exported to PDF:\n{pdf_path}")
            else:
                messagebox.showerror("Error", "Failed to export patient record to PDF")
        except ValueError:
            messagebox.showerror("Error", "ID must be a number")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to export patient record: {e}")

    def clear_form(self):
        try:
            for e in [self.e_id, self.e_name, self.e_age, self.e_phone, self.e_address,
                     self.e_occupation, self.e_diag, self.e_presc, self.e_doctor]:
                e.delete(0, "end")
            self.gender_cb.set("Male")
            self.current_image_blob = None
            self.patient_files = []
            self.photo_label.configure(image=None, text="No Photo")
        except Exception as e:
            print(f"Error clearing form: {e}")

    def load_all_patients(self):
        try:
            conn = db_connect()
            c = conn.cursor()

            for i in self.tree.get_children():
                self.tree.delete(i)

            c.execute("SELECT id, name, age, gender, phone, occupation, doctor, last_visit FROM patients ORDER BY id DESC")
            for row in c.fetchall():
                display_row = ["" if cell is None else cell for cell in row]
                self.tree.insert("", "end", values=display_row)
            conn.close()
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load patients: {e}")

    def search_patients(self):
        try:
            kw = self.search.get().strip()
            if not kw:
                self.load_all_patients()
                return

            conn = db_connect()
            c = conn.cursor()

            for i in self.tree.get_children():
                self.tree.delete(i)

            q = "SELECT id, name, age, gender, phone, occupation, doctor, last_visit FROM patients WHERE name LIKE ? OR phone LIKE ? OR doctor LIKE ? OR occupation LIKE ?"
            for row in c.execute(q, (f"%{kw}%", f"%{kw}%", f"%{kw}%", f"%{kw}%")):
                display_row = ["" if cell is None else cell for cell in row]
                self.tree.insert("", "end", values=display_row)
            conn.close()
        except Exception as e:
            messagebox.showerror("Error", f"Failed to search patients: {e}")

    def on_double(self, event):
        try:
            sel = self.tree.selection()
            if not sel:
                return
            pid = self.tree.item(sel[0], "values")[0]
            self.e_id.delete(0, "end")
            self.e_id.insert(0, str(pid))
            self.load_patient_by_id()
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load patient: {e}")

# ---------------- Visits View ----------------
class VisitsView:
    def __init__(self, parent):
        self.parent = parent

        parent.grid_columnconfigure(0, weight=1)
        parent.grid_rowconfigure(0, weight=1)

        frame = ctk.CTkFrame(parent, corner_radius=8, fg_color="#e2e8f0")
        frame.grid(row=0, column=0, padx=10, pady=10, sticky="nsew")
        frame.grid_columnconfigure(0, weight=1)
        frame.grid_rowconfigure(1, weight=1)

        ctk.CTkLabel(frame, text="Visit History",
                    font=ctk.CTkFont(size=18, weight="bold")).pack(pady=10)

        btn_frame = ctk.CTkFrame(frame, fg_color="transparent")
        btn_frame.pack(fill="x", padx=10, pady=5)

        ctk.CTkButton(btn_frame, text=icon_label("➕ Add Visit", "[+] Add Visit"), command=self.open_add,
                     fg_color="#27ae60", hover_color="#229954").pack(side="left", padx=5)
        ctk.CTkButton(btn_frame, text=icon_label("✏️ Edit Selected", "[Edit] Edit Selected"), command=self.open_edit,
                     fg_color="#f39c12", hover_color="#e67e22").pack(side="left", padx=5)
        ctk.CTkButton(btn_frame, text=icon_label("🗑️ Delete Selected", "[Del] Delete Selected"), fg_color="#e74c3c",
                     hover_color="#c0392b", command=self.delete_selected).pack(side="left", padx=5)

        filter_frame = ctk.CTkFrame(frame, fg_color="transparent")
        filter_frame.pack(fill="x", padx=10, pady=5)

        ctk.CTkLabel(filter_frame, text="Filter by Patient:").pack(side="left", padx=(0, 5))
        self.filter_var = ctk.StringVar()
        self.filter_cb = ctk.CTkComboBox(filter_frame, variable=self.filter_var, width=220)
        self.filter_cb.pack(side="left", padx=5)
        self.populate_filter()

        ctk.CTkButton(filter_frame, text=icon_label("🧹 Clear Filter", "[ ] Clear Filter"), command=self.clear_filter,
                     fg_color="#7f8c8d", hover_color="#95a5a6").pack(side="left", padx=5)
        ctk.CTkButton(filter_frame, text=icon_label("🔍 Apply Filter", "[?] Apply Filter"), command=self.apply_filter,
                     fg_color="#3498db", hover_color="#2980b9").pack(side="left", padx=5)

        table_frame = ctk.CTkFrame(frame, fg_color="transparent")
        table_frame.pack(fill="both", expand=True, padx=10, pady=10)
        table_frame.grid_columnconfigure(0, weight=1)
        table_frame.grid_rowconfigure(0, weight=1)

        cols = ("id", "patient", "date", "diagnosis", "prescription", "doctor", "price")
        self.tree = ttk.Treeview(table_frame, columns=cols, show="headings", height=20)

        self.tree.heading("id", text="Visit ID")
        self.tree.column("id", width=80, anchor="center")

        self.tree.heading("patient", text="Patient")
        self.tree.column("patient", width=150, anchor="w")

        self.tree.heading("date", text="Date")
        self.tree.column("date", width=120, anchor="center")

        self.tree.heading("diagnosis", text="Diagnosis")
        self.tree.column("diagnosis", width=150, anchor="w")

        self.tree.heading("prescription", text="Prescription")
        self.tree.column("prescription", width=150, anchor="w")

        self.tree.heading("doctor", text="Doctor")
        self.tree.column("doctor", width=120, anchor="w")

        self.tree.heading("price", text="Price ($)")
        self.tree.column("price", width=80, anchor="e")

        v_scrollbar = ctk.CTkScrollbar(table_frame, orientation="vertical", command=self.tree.yview)
        v_scrollbar.grid(row=0, column=1, sticky="ns")
        self.tree.configure(yscrollcommand=v_scrollbar.set)

        h_scrollbar = ctk.CTkScrollbar(table_frame, orientation="horizontal", command=self.tree.xview)
        h_scrollbar.grid(row=1, column=0, sticky="ew")
        self.tree.configure(xscrollcommand=h_scrollbar.set)

        self.tree.grid(row=0, column=0, sticky="nsew", padx=(0, 5))

        self.tree.bind("<Double-1>", lambda e: self.open_edit())

        self.load_visits()

    def populate_filter(self):
        try:
            conn = db_connect()
            c = conn.cursor()
            c.execute("SELECT id, name FROM patients ORDER BY name")
            rows = c.fetchall()
            conn.close()
            opts = ["All Patients"] + [f"{r[1]} (ID: {r[0]})" for r in rows]
            self.filter_cb.configure(values=opts)
            self.filter_cb.set("All Patients")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to populate filter: {e}")

    def load_visits(self):
        try:
            conn = db_connect()
            c = conn.cursor()

            for i in self.tree.get_children():
                self.tree.delete(i)

            q = """SELECT v.id, COALESCE(p.name, 'Unknown'), v.date, v.diagnosis, v.prescription, v.doctor, v.price
                   FROM visits v LEFT JOIN patients p ON v.patient_id = p.id
                   ORDER BY v.id DESC"""
            for row in c.execute(q):
                formatted_row = list(row)
                if formatted_row[6] is not None:
                    try:
                        formatted_row[6] = f"{float(formatted_row[6]):.2f}"
                    except (ValueError, TypeError):
                        formatted_row[6] = str(formatted_row[6])
                display_row = ["" if cell is None else cell for cell in formatted_row]
                self.tree.insert("", "end", values=display_row)
            conn.close()
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load visits: {e}")

    def clear_filter(self):
        try:
            self.filter_cb.set("All Patients")
            self.load_visits()
        except Exception as e:
            messagebox.showerror("Error", f"Failed to clear filter: {e}")

    def apply_filter(self):
        try:
            sel = self.filter_var.get()
            if not sel or sel == "All Patients":
                self.load_visits()
                return

            try:
                pid = int(sel.split("(ID: ")[1].split(")")[0])
            except (IndexError, ValueError):
                self.load_visits()
                return

            conn = db_connect()
            c = conn.cursor()

            for i in self.tree.get_children():
                self.tree.delete(i)

            q = """SELECT v.id, COALESCE(p.name, 'Unknown'), v.date, v.diagnosis, v.prescription, v.doctor, v.price
                   FROM visits v LEFT JOIN patients p ON v.patient_id = p.id
                   WHERE v.patient_id = ?
                   ORDER BY v.id DESC"""
            for row in c.execute(q, (pid,)):
                formatted_row = list(row)
                if formatted_row[6] is not None:
                    try:
                        formatted_row[6] = f"{float(formatted_row[6]):.2f}"
                    except (ValueError, TypeError):
                        formatted_row[6] = str(formatted_row[6])
                display_row = ["" if cell is None else cell for cell in formatted_row]
                self.tree.insert("", "end", values=display_row)
            conn.close()
        except Exception as e:
            messagebox.showerror("Error", f"Failed to apply filter: {e}")

    def open_add(self):
        try:
            # Guard when no patients exist
            conn = db_connect()
            c = conn.cursor()
            c.execute("SELECT COUNT(1) FROM patients")
            count = c.fetchone()[0]
            conn.close()
            if count == 0:
                messagebox.showerror("Error", "No patients found. Please add a patient first.")
                return
            self._open_popup(mode="add")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to open add visit dialog: {e}")

    def open_edit(self):
        try:
            sel = self.tree.selection()
            if not sel:
                messagebox.showerror("Error", "Select a visit to edit")
                return
            vid = self.tree.item(sel[0], "values")[0]
            self._open_popup(mode="edit", visit_id=vid)
        except Exception as e:
            messagebox.showerror("Error", f"Failed to open edit visit dialog: {e}")

    def _open_popup(self, mode="add", visit_id=None):
        try:
            popup = Toplevel()
            popup.title("Add Visit" if mode=="add" else "Edit Visit")
            popup.geometry("620x480")
            popup.resizable(False, False)

            ttk.Label(popup, text="Visit Information", font=("Arial", 16, "bold")).pack(pady=10)

            form_frame = ctk.CTkFrame(popup, corner_radius=8)
            form_frame.pack(fill="both", expand=True, padx=20, pady=10)

            # Patient selection
            ttk.Label(form_frame, text="Patient:").place(x=20, y=20)
            conn = db_connect()
            c = conn.cursor()
            c.execute("SELECT id, name FROM patients ORDER BY name")
            patients = c.fetchall()
            conn.close()

            opts = [f"{r[1]} (ID: {r[0]})" for r in patients]
            patient_var = ctk.StringVar()
            patient_cb = ttk.Combobox(form_frame, values=opts, textvariable=patient_var, width=50, state="readonly")
            patient_cb.place(x=120, y=20)

            if not opts:
                patient_cb.configure(state="disabled")
                ttk.Label(form_frame, text="No patients found. Add a patient first.", foreground="red").place(x=120, y=50)
            else:
                patient_var.set(opts[0])

            # Date
            ttk.Label(form_frame, text="Date (YYYY-MM-DD HH:MM):").place(x=20, y=60)
            date_e = ttk.Entry(form_frame, width=28)
            date_e.place(x=240, y=60)
            date_e.insert(0, datetime.now().strftime("%Y-%m-%d %H:%M"))

            # Diagnosis
            ttk.Label(form_frame, text="Diagnosis:").place(x=20, y=100)
            diag_e = ttk.Entry(form_frame, width=55)
            diag_e.place(x=120, y=100)

            # Prescription
            ttk.Label(form_frame, text="Prescription:").place(x=20, y=140)
            presc_e = ttk.Entry(form_frame, width=55)
            presc_e.place(x=120, y=140)

            # Doctor
            ttk.Label(form_frame, text="Doctor:").place(x=20, y=180)
            doc_e = ttk.Entry(form_frame, width=55)
            doc_e.place(x=120, y=180)

            # Price
            ttk.Label(form_frame, text="Price ($):").place(x=20, y=220)
            price_e = ttk.Entry(form_frame, width=20)
            price_e.place(x=120, y=220)

            # If editing, load data
            if mode == "edit" and visit_id:
                try:
                    conn = db_connect()
                    c = conn.cursor()
                    c.execute("SELECT id, patient_id, date, diagnosis, prescription, doctor, price FROM visits WHERE id=?", (visit_id,))
                    v = c.fetchone()
                    conn.close()
                    if v:
                        _, patient_id, date, diagnosis, prescription, doctor, price = v

                        # match patient
                        matched = False
                        for o in opts:
                            if o.endswith(f"(ID: {patient_id})"):
                                patient_var.set(o)
                                matched = True
                                break
                        if not matched and opts:
                            patient_var.set(opts[0])

                        date_e.delete(0, "end")
                        date_e.insert(0, date or "")

                        diag_e.delete(0, "end")
                        diag_e.insert(0, diagnosis or "")

                        presc_e.delete(0, "end")
                        presc_e.insert(0, prescription or "")

                        doc_e.delete(0, "end")
                        doc_e.insert(0, doctor or "")

                        price_e.delete(0, "end")
                        price_e.insert(0, str(price) if price is not None else "")
                except Exception as e:
                    messagebox.showerror("Error", f"Failed to load visit data: {e}")
                    popup.destroy()
                    return

            def parse_patient_id(sel_text):
                try:
                    return int(sel_text.split("(ID: ")[1].split(")")[0])
                except Exception:
                    return None

            def save_visit():
                try:
                    if not opts:
                        messagebox.showerror("Error", "No patients available. Please add a patient first.")
                        return

                    sel = patient_var.get().strip()
                    pid = parse_patient_id(sel)
                    if pid is None:
                        messagebox.showerror("Error", "Select a valid patient")
                        return

                    dt = date_e.get().strip()
                    if not dt:
                        messagebox.showerror("Error", "Date is required")
                        return
                    try:
                        datetime.strptime(dt, "%Y-%m-%d %H:%M")
                    except ValueError:
                        messagebox.showerror("Error", "Date must be in format YYYY-MM-DD HH:MM")
                        return

                    diag = diag_e.get().strip()
                    presc = presc_e.get().strip()
                    doc = doc_e.get().strip()

                    price_str = price_e.get().strip()
                    price = 0.0
                    if price_str:
                        try:
                            price = float(price_str)
                            if price < 0:
                                messagebox.showerror("Error", "Price cannot be negative")
                                return
                        except ValueError:
                            messagebox.showerror("Error", "Price must be a number")
                            return

                    conn = db_connect()
                    c = conn.cursor()

                    if mode == "add":
                        c.execute('''INSERT INTO visits (patient_id, date, diagnosis, prescription, doctor, price)
                                     VALUES (?, ?, ?, ?, ?, ?)''', (pid, dt, diag, presc, doc, price))
                        c.execute("UPDATE patients SET last_visit=? WHERE id=?", (dt, pid))
                    else:
                        c.execute('''UPDATE visits SET patient_id=?, date=?, diagnosis=?, prescription=?, doctor=?, price=? WHERE id=?''',
                                  (pid, dt, diag, presc, doc, price, visit_id))
                        c.execute("UPDATE patients SET last_visit=? WHERE id=?", (dt, pid))

                    conn.commit()
                    conn.close()
                    messagebox.showinfo("Success", "Visit saved successfully")
                    popup.destroy()
                    self.load_visits()
                    self.populate_filter()
                except Exception as e:
                    messagebox.showerror("Error", f"Failed to save visit: {e}")

            ttk.Button(form_frame, text="Save Visit", command=save_visit).place(x=260, y=340)
        except Exception as e:
            messagebox.showerror("Error", f"Failed to create visit dialog: {e}")

    def delete_selected(self):
        try:
            sel = self.tree.selection()
            if not sel:
                messagebox.showerror("Error", "Select a visit to delete")
                return
            vid = self.tree.item(sel[0], "values")[0]
            if messagebox.askyesno("Confirm Delete", "Are you sure you want to delete this visit?"):
                conn = db_connect()
                c = conn.cursor()
                c.execute("DELETE FROM visits WHERE id=?", (vid,))
                conn.commit()
                conn.close()
                messagebox.showinfo("Success", "Visit deleted successfully")
                self.load_visits()
                self.populate_filter()
        except Exception as e:
            messagebox.showerror("Error", f"Failed to delete visit: {e}")

# ---------------- Users View ----------------
class UsersView:
    def __init__(self, parent):
        parent.grid_columnconfigure(0, weight=1)
        parent.grid_rowconfigure(0, weight=1)

        frame = ctk.CTkFrame(parent, corner_radius=8, fg_color="#e2e8f0")
        frame.grid(row=0, column=0, padx=10, pady=10, sticky="nsew")
        frame.grid_columnconfigure(0, weight=1)
        frame.grid_rowconfigure(1, weight=1)

        ctk.CTkLabel(frame, text="User Management (Admin)",
                    font=ctk.CTkFont(size=18, weight="bold")).pack(pady=10)

        form_frame = ctk.CTkFrame(frame, corner_radius=8)
        form_frame.pack(fill="x", padx=10, pady=5)

        ctk.CTkLabel(form_frame, text="Add New User",
                    font=ctk.CTkFont(size=14, weight="bold")).pack(pady=10)

        ctk.CTkLabel(form_frame, text="Username:").pack(anchor="w", padx=10)
        self.u_name = ctk.CTkEntry(form_frame, placeholder_text="Enter username")
        self.u_name.pack(fill="x", padx=10, pady=5)

        ctk.CTkLabel(form_frame, text="Password:").pack(anchor="w", padx=10)
        self.u_pass = ctk.CTkEntry(form_frame, placeholder_text="Enter password", show="*")
        self.u_pass.pack(fill="x", padx=10, pady=5)

        ctk.CTkLabel(form_frame, text="Role:").pack(anchor="w", padx=10)
        self.u_role = ctk.CTkEntry(form_frame, placeholder_text="Enter role (e.g., Admin, Doctor)")
        self.u_role.pack(fill="x", padx=10, pady=5)

        ctk.CTkButton(form_frame, text="Add User", command=self.add_user,
                     fg_color="#27ae60", hover_color="#229954").pack(pady=10)

        ctk.CTkLabel(form_frame, text="Default admin: abdo / 202300488",
                    font=ctk.CTkFont(size=10)).pack(pady=5)

        table_frame = ctk.CTkFrame(frame, fg_color="transparent")
        table_frame.pack(fill="both", expand=True, padx=10, pady=10)
        table_frame.grid_columnconfigure(0, weight=1)
        table_frame.grid_rowconfigure(0, weight=1)

        self.tree = ttk.Treeview(table_frame, columns=("id", "username", "role"),
                                show="headings", height=15)
        self.tree.heading("id", text="ID")
        self.tree.column("id", width=50, anchor="center")
        self.tree.heading("username", text="Username")
        self.tree.column("username", width=200, anchor="w")
        self.tree.heading("role", text="Role")
        self.tree.column("role", width=150, anchor="w")

        v_scrollbar = ctk.CTkScrollbar(table_frame, orientation="vertical", command=self.tree.yview)
        v_scrollbar.grid(row=0, column=1, sticky="ns")
        self.tree.configure(yscrollcommand=v_scrollbar.set)

        h_scrollbar = ctk.CTkScrollbar(table_frame, orientation="horizontal", command=self.tree.xview)
        h_scrollbar.grid(row=1, column=0, sticky="ew")
        self.tree.configure(xscrollcommand=h_scrollbar.set)

        self.tree.grid(row=0, column=0, sticky="nsew", padx=(0, 5))

        action_frame = ctk.CTkFrame(frame, fg_color="transparent")
        action_frame.pack(fill="x", padx=10, pady=10)

        ctk.CTkButton(action_frame, text=icon_label("🗑️ Delete Selected", "[Del] Delete Selected"), fg_color="#e74c3c",
                     hover_color="#c0392b", command=self.delete_selected).pack(side="left", padx=5)
        ctk.CTkButton(action_frame, text=icon_label("🔄 Refresh", "[R] Refresh"), command=self.load_users).pack(side="left", padx=5)

        self.load_users()

    def add_user(self):
        try:
            uname = self.u_name.get().strip()
            pwd = self.u_pass.get().strip()
            role = self.u_role.get().strip()

            if not uname or not pwd:
                messagebox.showerror("Error", "Username and password are required")
                return

            if not role:
                messagebox.showerror("Error", "Role is required")
                return

            conn = db_connect()
            c = conn.cursor()
            c.execute("INSERT INTO users (username, password, role) VALUES (?, ?, ?)", (uname, pwd, role))
            conn.commit()
            conn.close()
            messagebox.showinfo("Success", "User added successfully")
            self.u_name.delete(0, "end")
            self.u_pass.delete(0, "end")
            self.u_role.delete(0, "end")
            self.load_users()
        except sqlite3.IntegrityError:
            messagebox.showerror("Error", "Username already exists")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to add user: {e}")

    def load_users(self):
        try:
            conn = db_connect()
            c = conn.cursor()

            for i in self.tree.get_children():
                self.tree.delete(i)

            for row in c.execute("SELECT id, username, role FROM users"):
                display_row = ["" if cell is None else cell for cell in row]
                self.tree.insert("", "end", values=display_row)
            conn.close()
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load users: {e}")

    def delete_selected(self):
        try:
            sel = self.tree.selection()
            if not sel:
                messagebox.showerror("Error", "Select a user to delete")
                return
            uid = self.tree.item(sel[0], "values")[0]

            conn = db_connect()
            c = conn.cursor()
            c.execute("SELECT username FROM users WHERE id=?", (uid,))
            row = c.fetchone()
            if row and row[0] == "abdo":
                conn.close()
                messagebox.showerror("Error", "Cannot delete the default admin user")
                return

            if messagebox.askyesno("Confirm Delete", "Are you sure you want to delete this user?"):
                c.execute("DELETE FROM users WHERE id=?", (uid,))
                conn.commit()
                conn.close()
                messagebox.showinfo("Success", "User deleted successfully")
                self.load_users()
        except Exception as e:
            messagebox.showerror("Error", f"Failed to delete user: {e}")

if __name__ == "__main__":
    try:
        LoginWindow().mainloop()
    except Exception as e:
        print(f"Fatal error: {e}")
        traceback.print_exc()
